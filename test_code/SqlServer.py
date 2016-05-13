# -*- encoding:utf8 -*-
import pyodbc
import openpyxl
from WindPy import *
import os
w.start()
conn_info = 'DRIVER={SQL Server};DATABASE=stock;SERVER=DESKTOP-SDEEIE3\CQY;'
mssql_conn = pyodbc.connect(conn_info)
mssql_cur = mssql_conn.cursor()
code = '600111.SH'
date = '20140819'
sql = "Select top 1000 * from [stock].[dbo].[vw_Transactions] where WindCode = '" + code + "' and BizDate = '" + date + "'"
mssql_cur.execute(sql)
row = mssql_cur.fetchall()
for each in row:
    print each
print 123

wb = openpyxl.load_workbook('2.xlsx')
ws = wb.get_sheet_by_name('Sheet1')
for i in range(2,18):
    code = ws.cell(row = i, column = 1).value

    start1 = str(ws.cell(row = i, column = 2).value)[0:10]
    start2 = w.tdaysoffset(-21, start1).Data[0][0].isoformat()[0:10]
    start = start2[0:4] + start2[5:7] + start2[8:10]

    end1 = str(ws.cell(row = i, column = 3).value)[0:10]
    end2 = w.tdaysoffset(20, end1).Data[0][0].isoformat()[0:10]
    end = end2[0:4] + end2[5:7] + end2[8:10]

    next_day1 = start2
    while cmp(next_day1,end2):
        next_day1 = w.tdaysoffset(1, next_day1).Data[0][0].isoformat()[0:10]
        next_day = next_day1[0:4] + next_day1[5:7] + next_day1[8:10]
        sql = "Select * from [stock].[dbo].[vw_Transactions] where WindCode = '" + code + "' and BizDate = '" + next_day + "'"
        mssql_cur.execute(sql)
        row = mssql_cur.fetchall()
        name = "E:/transactions/" + code + "/" + next_day + ".csv"
        f = open(name,"a+")
        for each in row:
            a = str(each[1].encode('utf8'))
            line = str(each[0]) + "," + a
            line = line + "," + str(each[2]) + "," + str(each[3]) + "," + str(each[4]) + "," + str(each[5]) + "," + str(each[6]) + "," + str(each[7]) + "," + str(each[8]) + "," + str(each[9]) + "," + "," + str(each[12]) + "," + str(each[13]) + "\n"
            f.writelines(str(line))
        f.close()
        print code + next_day
    # cur_dir = 'E:\mypython\chenqiuyuan\\test_code\\transactions'
    # folder_name = code
    # folder_path = cur_dir + "\\" + code
    # if os.path.isdir(cur_dir):
    #     if not os.path.exists(folder_path):
    #         os.mkdir(os.path.join(cur_dir, folder_name))
    print 123